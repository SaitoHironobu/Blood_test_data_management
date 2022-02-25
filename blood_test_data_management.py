import PySimpleGUI as sg
import pandas as pd
from pathlib import Path
import random
import openpyxl


sg.SetOptions(text_justification='right')
layout = [
    [sg.Text('血液データ登録フォーム',font=('小塚ゴシック pro B',16), pad=(50,0))],
    [sg.Text('検査日',size=(15,1)), sg.InputText(key='-Day-',size=(18,1))],
    [sg.Text('LDLコレステロール',size=(15,1)), sg.InputText(key='-Ldl-', size=(10,1))],
    [sg.Text('中性脂肪(TG)',size=(15,1)), sg.InputText(key='-Tg-', size=(10,1))],
    [sg.Text('尿素窒素(BUN)',size=(15,1)), sg.InputText(key='-Bun-', size=(10,1))],
    [sg.Text('クレアチニン',size=(15,1)), sg.InputText(key='-Cr-', size=(10,1))],
    [sg.Text('尿酸(UA)',size=(15,1)), sg.InputText(key='-Ua-', size=(10,1))],
    [sg.Text('血糖(グレコース)',size=(15,1)), sg.InputText(key='-Bls-', size=(10,1))],
    [sg.Text('HbA1c',size=(15,1)), sg.InputText(key='-Hba1c-', size=(10,1))],
    [sg.Text('体重',size=(15,1)), sg.InputText(key='-Bw-', size=(10,1))],
    [sg.Text('性別',size=(15,1)), sg.Radio('男', 'sex', default= True, key='-Male-'),
    sg.Radio('女', 'sex',  key='-Female-')],
    [sg.Text('年齢',size=(15,1)), sg.Spin([i for i in range(20,90)], initial_value=35, key='-Age-')],
    [sg.Button('新規登録', key='-Register-', pad=(120,0), disabled=False)]
]

window = sg.Window('血液データ登録フォーム', layout)

def get_wgt_values():
    day = value['-Day-']
    ldl = value['-Ldl-']
    tg = value['-Tg-']
    bun = value['-Bun-']
    cr = value['-Cr-']
    ua = value['-Ua-']
    bls = value['-Bls-']
    hba1c = value['-Hba1c-']
    bw = value['-Bw-']
    if value['-Male-'] is True:
        sex = '男'
    else:
        sex = '女'
        
    age = value['-Age-']
        
        
    data = [day, ldl, tg, bun, cr, ua, bls, hba1c, bw, sex, age]
    
    return data

def insert_to_db(data):
    try:
        wb = openpyxl.load_workbook('血液検査データ.xlsx')
        ws = wb['Sheet1']
        last_row = ws.max_row + 1
        for colm in range(len(data)):
            ws.cell(row=last_row, column=colm + 1, value=data[colm])
        wb.save('血液検査データ.xlsx')
        status = 'succeed'
    except:
        status = None
    return status

def wgt_clear():
    window.find_element('-Day-').Update('')
    window.find_element('-Ldl-').Update('')
    window.find_element('-Tg-').Update('')
    window.find_element('-Bun-').Update('')
    window.find_element('-Cr-').Update('')
    window.find_element('-Ua-').Update('')
    window.find_element('-Bls-').Update('')
    window.find_element('-Hba1c-').Update('')
    window.find_element('-Bw-').Update('')
    

while True:
    event, value = window.read() 
    if event == '-Register-':
        blood_test_data = get_wgt_values()
        
        result = insert_to_db(blood_test_data)
        
        if result == 'succeed':
            sg.popup('追記が成功しました.')
            wgt_clear()
        else:
            sg.popup('追記が失敗しました。')

        
    elif event is None:
        break
    
window.close()